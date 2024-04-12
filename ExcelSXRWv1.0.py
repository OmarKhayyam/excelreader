import openpyxl
import boto3
import json
import os

classification_bedrock_models_map = {
    'us-east-1': 'amazon.titan-text-lite-v1',
    'us-west-2': 'anthropic.claude-instant-v1',
    'ap-southeast-1': 'anthropic.claude-instant-v1',
    'ap-southeast-2': 'amazon.titan-text-lite-v1',
    'ap-northeast-1': 'amazon.titan-text-express-v1',
    'eu-central-1': 'amazon.titan-text-express-v1',
    'eu-west-3': 'amazon.titan-text-lite-v1'
}

class ExcelSXRW:
    def __init__(self,filepath,numrowstosearch=10,numcolumnstosearch=10,sheets_to_process=[]) -> None:
        """We get the path and info whether there is a header row."""
        self.path = filepath
        self.query_locations = list()
        self.excelfile = openpyxl.load_workbook(self.path)
        if len(sheets_to_process) == 0:
            self.sheets_to_process = [] ## When 0 sheets are provided, all sheets to be processed
            for sheet in self.excelfile.sheetnames:
                self.sheets_to_process.append(self.excelfile.sheetnames.index(sheet))
        else :
            self.sheets_to_process = sheets_to_process ## Expecting sheets to start from any number, list of numbers
        self.region_name = os.environ['AWS_REGION']
        self.classification_model_id = classification_bedrock_models_map[self.region_name]
        self.bedrock_client = boto3.client('bedrock-runtime')
        self.rowstosearch = numrowstosearch # These are the rows we will search for the first query
        self.colstosearch = numcolumnstosearch # These are the columns we will search for the first query

    def _anthropic_claude_instant_prompt(self,cell_value) -> str:
        """Prompting Anthropic Claude Instant model for a True or False response"""
        parameterised_text = """Human: Read the following statement, and classify whether or not the statement is asking for information. Respond in one word, use the word YES if it is asking for information or NO. Statement: {actual_query}. Assistant:""".format(actual_query=cell_value)
        return parameterised_text 

    def _amazon_titan_text_lite_prompt(self,cell_value) -> str:
        """UNTESTED: Prompting Amazon Titan Text model for a True or False response"""
        parameterised_text = """Respond in terms of True or False whether the following statement should be a question. What are the major conflict, setting, and theme in The Hitchhiker's Guide to the Galaxy.{actual_query}.""".format(actual_query=cell_value)
        return parameterised_text
    
    def _amazon_titan_text_express_prompt(self, cell_value) -> str:
        """UNTESTED: Prompting Amazon Titan Text model for a True or False response"""
        parameterised_text = """Simply respond in terms of True or False whether the following statement should be a question, ignore punctuation. Only respond in a single word.\nStatement: {actual_query}.""".format(actual_query=cell_value)
        return parameterised_text
    
    def _get_prompt(self,cell_value) -> str:
        """Get the right kind of prompt for the right kind of model"""
        if self.region_name in ['ap-southeast-1','us-west-2']:
            prompt = self._anthropic_claude_instant_prompt(cell_value)
        elif self.region_name in ['us-east-1','eu-west-3']:
            prompt = self._amazon_titan_text_lite_prompt(cell_value)
        else: ## For all other regions
            prompt = self._amazon_titan_text_express_prompt(cell_value)
        return prompt

    def locate_queries(self) -> list:
        """Get the queries from the excel file."""
        ## Get the queries from the excel file
        ## Return a list of tuples, each tuple is the sheet index, row index, column index
        ## For example, (0, 2, 3) means the query is in the first sheet, row 2, column 3
        self.query_locations = self._find_first_query_row().copy()

    def get_sheets(self) -> list:
        """Get the list of sheets to be processed."""
        return self.sheets_to_process

    def get_queries_in_sheet(self, sheet_index) -> list:
        """Get the queries in a sheet."""
        ## Get the queries in a sheet
        ## Return a list of tuples, each tuple is the sheet index, row index, column index
        ## For example, (0, 2, 3) means the query is in the first sheet, row 2, column 3
        queries = list()
        for query in self.query_locations:
            if query[0] == sheet_index:
                self.active_sheet = self.excelfile.worksheets[sheet_index]
                queries.append(self.active_sheet.cell(row = query[1], column = query[2]).value)
                row = query[1]
                col = query[2]
                while self.active_sheet.cell(row = row+1, column = query[2]).value != None:
                    queries.append(self.active_sheet.cell(row = row+1, column = query[2]).value)
                    row += 1
        return queries

    def _is_cell_value_a_query(self, cell_value) -> bool:
        """Checks if the cell value is a query."""
        ## Take this data in the cell_value and check if it is a query, 
        ## return True if it is a query.
        #body = json.dumps({"prompt": "Human: {prmpt}\nAssistant:".format(prmpt=self._get_prompt(cell_value)), "max_tokens_to_sample": 100})
        body = json.dumps({"prompt": "{prmpt}".format(prmpt=self._get_prompt(cell_value)), "max_tokens_to_sample": 100})
        #print("**THE BODY IS {}".format(body))
        modelId = self.classification_model_id
        #print("**THE MODEL ID is {}".format(modelId))
        accept = 'application/json'
        content_type = 'application/json'
        response = self.bedrock_client.invoke_model(body=body, modelId=modelId, accept=accept, contentType=content_type)
        response_json = json.loads(response['body'].read().decode('utf-8'))
        #print(response_json['completion'])
        if "NO" in response_json['completion']:
            return False
        return True

    def _find_first_query_row(self) -> list:
        """Finds the first row of the queries and its column."""
        found_cell,query_cells = False,list()
        for sheet_index in self.sheets_to_process:
            self.active_sheet = self.excelfile.worksheets[sheet_index]
            for j in range(1,self.colstosearch): ## for every column
                for i in range(1,self.rowstosearch): ## for every row
                    cell_value = self.active_sheet.cell(row = i, column = j).value
                    #print("cell_value: {}".format(cell_value))
                    #print("Checking row: {} and column: {}".format(i,j))
                    if cell_value == None:
                        continue
                    if self._is_cell_value_a_query(cell_value) == True:
                        query_cells.append((sheet_index,i, j))
                        found_cell = True
                        break
                if found_cell == True:
                    found_cell = False
                    break
        return query_cells ## Could not find a question in all the sheets.

### This code has only been tested with us-west-2 region. ###
### The main() is only being used as a driver. ###
if __name__ == "__main__":
    test = ExcelSXRW("Documents/Explorations/QnA_Samples/Queries_and_Responses.xlsx")
    #test._is_cell_value_a_query("What are the major conflict, setting, and theme in The Hitchhiker's Guide to the Galaxy")
    #test._is_cell_value_a_query("The Hitchhikers Guide to the Galaxy is an awesome book!")
    test.locate_queries()
    for sheet in test.get_sheets():
        print(test.get_queries_in_sheet(sheet))  